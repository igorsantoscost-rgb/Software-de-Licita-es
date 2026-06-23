"""Importador da base CAPAG (Tesouro Nacional) para o banco local.

Baixa o arquivo MAIS RECENTE de estados (CSV) e de municipios (XLSX) do portal
Tesouro Transparente, via API CKAN, e popula as tabelas capag_estados e
capag_municipios.

Como rodar manualmente (dentro do container, na VPS):
    docker compose exec app python -m app.capag_import

Tambem e chamado pelo botao "Atualizar base CAPAG" dentro do sistema.
"""
import io
import os
import csv
import tempfile
from datetime import datetime

import requests

from app.capag import normalizar

CKAN_BASE = "https://www.tesourotransparente.gov.br/ckan/api/3/action/package_show"
PKG_ESTADOS = "capag-estados"
PKG_MUNICIPIOS = "capag-municipios"
HEADERS = {"User-Agent": "BIDFY-CAPAG/1.0 (consultoria de licitacoes)"}


def _resource_mais_recente(package_id, formatos):
    """Consulta a API CKAN e devolve a URL do recurso mais recente cujo formato
    esteja em `formatos`, ignorando arquivos de metadados."""
    r = requests.get(CKAN_BASE, params={"id": package_id}, headers=HEADERS, timeout=60)
    r.raise_for_status()
    recursos = r.json()["result"]["resources"]
    candidatos = []
    for rec in recursos:
        fmt = (rec.get("format") or "").lower().strip()
        nome = (rec.get("name") or "").lower()
        url = rec.get("url") or ""
        if "metadado" in nome:
            continue
        casa_formato = fmt in formatos or any(url.lower().endswith("." + f) for f in formatos)
        if not casa_formato:
            continue
        data = rec.get("created") or rec.get("last_modified") or ""
        candidatos.append((data, url))
    if not candidatos:
        raise RuntimeError(f"Nenhum recurso encontrado para {package_id}")
    candidatos.sort(reverse=True)  # datas ISO ordenam corretamente; mais recente primeiro
    return candidatos[0][1]


def _baixar(url):
    r = requests.get(url, headers=HEADERS, timeout=300)
    r.raise_for_status()
    return r.content


def _achar_indice(cabecalho, *grupos):
    """Acha o indice da coluna cujo nome (sem acento) contem TODAS as palavras
    de algum dos grupos. Ex.: _achar_indice(cab, ['classificacao','capag'])."""
    norm = [normalizar(c) for c in cabecalho]
    for grupo in grupos:
        for i, nome in enumerate(norm):
            if all(p in nome for p in grupo):
                return i
    return None


def importar_estados():
    from app import db
    from app.models import CapagEstado

    url = _resource_mais_recente(PKG_ESTADOS, {"csv"})
    bruto = _baixar(url)
    texto = None
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            texto = bruto.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if texto is None:
        raise RuntimeError("Nao consegui decodificar o CSV de estados")

    primeira = texto.splitlines()[0] if texto.splitlines() else ""
    sep = ";" if primeira.count(";") >= primeira.count(",") else ","
    linhas = list(csv.reader(io.StringIO(texto), delimiter=sep))
    if not linhas:
        return 0

    cab = linhas[0]
    i_uf = _achar_indice(cab, ["uf"], ["sigla"], ["estado"])
    i_classe = _achar_indice(cab, ["classificacao", "capag"], ["classificacao"], ["nota", "capag"])
    if i_uf is None or i_classe is None:
        raise RuntimeError(f"Colunas de estados nao reconhecidas: {cab}")

    ref = f"Tesouro Transparente, atualizado em {datetime.now():%d/%m/%Y}"
    CapagEstado.query.delete()
    n = 0
    for row in linhas[1:]:
        if len(row) <= max(i_uf, i_classe):
            continue
        uf = (row[i_uf] or "").strip().upper()
        if len(uf) != 2:
            continue
        classe = (row[i_classe] or "").strip().upper()
        db.session.add(CapagEstado(uf=uf, classificacao=classe, referencia=ref))
        n += 1
    db.session.commit()
    return n


def importar_municipios():
    from app import db
    from app.models import CapagMunicipio
    from openpyxl import load_workbook

    url = _resource_mais_recente(PKG_MUNICIPIOS, {"xlsx"})
    bruto = _baixar(url)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(bruto)
        caminho = tmp.name

    try:
        wb = load_workbook(caminho, read_only=True, data_only=True)
        ws = wb.active
        linhas = ws.iter_rows(values_only=True)

        # Acha a linha de cabecalho (algumas planilhas tem linhas de titulo antes)
        cab = None
        tentativas = 0
        for row in linhas:
            tentativas += 1
            valores = [str(c) if c is not None else "" for c in row]
            if (_achar_indice(valores, ["uf"], ["sigla"]) is not None and
                    _achar_indice(valores, ["classificacao", "capag"], ["classificacao"], ["capag"]) is not None):
                cab = valores
                break
            if tentativas > 15:
                break
        if cab is None:
            raise RuntimeError("Cabecalho da planilha de municipios nao encontrado")

        i_uf = _achar_indice(cab, ["uf"], ["sigla"])
        i_classe = _achar_indice(cab, ["classificacao", "capag"], ["classificacao"], ["capag"])
        i_nome = _achar_indice(cab, ["municipio"], ["ente"], ["instituicao"], ["nome"])
        i_ibge = _achar_indice(cab, ["ibge"], ["cod", "mun"])
        if i_uf is None or i_classe is None or i_nome is None:
            raise RuntimeError(f"Colunas de municipios nao reconhecidas: {cab}")

        ref = f"Tesouro Transparente, atualizado em {datetime.now():%d/%m/%Y}"
        CapagMunicipio.query.delete()
        db.session.commit()

        maxi = max(i_uf, i_classe, i_nome, i_ibge if i_ibge is not None else 0)
        n, lote = 0, 0
        for row in linhas:  # continua do ponto seguinte ao cabecalho
            if row is None or len(row) <= maxi:
                continue
            uf = str(row[i_uf]).strip().upper() if row[i_uf] is not None else ""
            nome = str(row[i_nome]).strip() if row[i_nome] is not None else ""
            if len(uf) != 2 or not nome:
                continue
            classe = str(row[i_classe]).strip().upper() if row[i_classe] is not None else ""
            ibge = str(row[i_ibge]).strip() if (i_ibge is not None and row[i_ibge] is not None) else None
            db.session.add(CapagMunicipio(
                cod_ibge=ibge, uf=uf, nome=nome,
                nome_normalizado=normalizar(nome),
                classificacao=classe, referencia=ref,
            ))
            n += 1
            lote += 1
            if lote >= 1000:
                db.session.commit()
                lote = 0
        db.session.commit()
        wb.close()
        return n
    finally:
        try:
            os.remove(caminho)
        except OSError:
            pass


def importar_tudo():
    """Importa estados e municipios. Retorna dict com contagens e erros."""
    resultado = {"estados": 0, "municipios": 0, "erros": []}
    try:
        resultado["estados"] = importar_estados()
    except Exception as e:
        resultado["erros"].append(f"Estados: {e}")
    try:
        resultado["municipios"] = importar_municipios()
    except Exception as e:
        resultado["erros"].append(f"Municipios: {e}")
    return resultado


if __name__ == "__main__":
    from app import create_app
    app = create_app()
    with app.app_context():
        print("Importando base CAPAG do Tesouro Nacional...")
        res = importar_tudo()
        print(f"  Estados:    {res['estados']}")
        print(f"  Municipios: {res['municipios']}")
        if res["erros"]:
            print("  ERROS:")
            for e in res["erros"]:
                print("   -", e)
        else:
            print("  OK!")
